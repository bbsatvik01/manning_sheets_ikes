"""
Manning Chart Web Application
============================

This script provides a simple web interface for generating Manning Chart
workbooks from staff scheduling spreadsheets. It supports multiple locations
(e.g., "Ikes" and "Southside") with configurable station mappings.

"""

import os
import re
import sys
import threading
import logging
import webbrowser
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from flask import (
    Flask,
    request,
    redirect,
    url_for,
    render_template_string,
    send_from_directory,
    abort,
    flash,
    get_flashed_messages,
)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "Required dependency openpyxl is missing. Install with 'pip install openpyxl'."
    )


###############################
# Configuration and utilities #
###############################

# Determine base directory for read/write data and locate bundled resources
if getattr(sys, "frozen", False):
    EXEC_DIR = os.path.dirname(sys.executable)
    RESOURCE_DIR = getattr(sys, "_MEIPASS", EXEC_DIR)
else:
    EXEC_DIR = os.path.dirname(os.path.abspath(__file__))
    RESOURCE_DIR = EXEC_DIR

BASE_DIR = EXEC_DIR
INPUT_DIR = os.path.join(BASE_DIR, "input_my_staff_schedule")
OUTPUT_DIR = os.path.join(BASE_DIR, "manning_sheets")
LOG_DIR = os.path.join(BASE_DIR, "logs")
STATIC_ROOT = os.path.join(RESOURCE_DIR, "static")
if not os.path.isdir(STATIC_ROOT):
    STATIC_ROOT = os.path.join(BASE_DIR, "static")

# Ensure necessary directories exist
os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(LOG_DIR, exist_ok=True)

# Track outputs generated during this runtime (latest batch only)
CURRENT_OUTPUTS: List[str] = []

# Locations configuration
LOCATIONS = {
    "ikes": {"name": "Ikes", "mapping_needed": False},
    "southside": {"name": "Southside", "mapping_needed": True},
}


from mappings import SOUTHSIDE_MAPPING, IKES_MAPPING, SOUTHSIDE_KEYWORDS, IKES_KEYWORDS

def parse_time(time_str: str) -> Optional[float]:
    """Convert a 12‑hour time string into a floating point hour.

    Example: "06:00 AM" -> 6.0, "02:30 PM" -> 14.5.
    Returns None if parsing fails.
    """
    time_str = time_str.strip()
    match = re.match(r"(\d{1,2}):(\d{2})\s*(AM|PM)", time_str, re.IGNORECASE)
    if not match:
        return None
    hour, minute, meridiem = int(match.group(1)), int(match.group(2)), match.group(3).upper()
    if meridiem == "PM" and hour != 12:
        hour += 12
    if meridiem == "AM" and hour == 12:
        hour = 0
    return hour + minute / 60.0


# Additional mappings not in the Excel file
FALLBACK_MAPPINGS = {
    "am pasta": "LITTLE ITALY",
}

# Roles to strictly ignore (not count as missing assignments)
IGNORED_ROLES = {
    "scheduled elsewhere",
}


def validate_file_location(file_path: str, location: str) -> bool:
    """Validate that the uploaded file matches the expected location."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        header_val = str(ws['A1'].value or "")
        
        if location == "ikes":
            return "GMU DH-Ike" in header_val
        elif location == "southside":
            return "GMU DH-Southside" in header_val
            
        return False
    except Exception as e:
        logging.error(f"Validation error for {file_path}: {e}")
        return False


def get_category(role: str, location: str) -> Optional[str]:
    """Map a job role to a Manning Chart category (station) based on location.
    
    1. Exact match (case-insensitive) against dictionary.
    2. Fallback exact matches.
    3. Fuzzy keyword match.
    """
    if not role:
        return None
    role_clean = role.strip().replace("\n", "").strip()
    role_lower = role_clean.lower()

    if location == "southside":
        # 1. Exact lookup
        if role_lower in SOUTHSIDE_MAPPING:
            return SOUTHSIDE_MAPPING[role_lower]
        if role_lower in FALLBACK_MAPPINGS:
            return FALLBACK_MAPPINGS[role_lower]
            
        # 2. Fuzzy/Keyword lookup
        for keyword, station in SOUTHSIDE_KEYWORDS:
            if keyword in role_lower:
                return station
                
        return None

    elif location == "ikes":
        # 1. Exact lookup
        if role_lower in IKES_MAPPING:
            return IKES_MAPPING[role_lower]
            
        # 2. Fuzzy/Keyword lookup
        for keyword, station in IKES_KEYWORDS:
            if keyword in role_lower:
                return station

        return None
    
    return None


def get_stations_layout(location: str) -> List[List[str]]:
    """Return the grid layout of stations for the Manning Sheet."""
    if location == "southside":
        # 5-column layout for Southside
        known_stations = sorted(list(set(SOUTHSIDE_MAPPING.values())))
        # Filter out None or empty
        known_stations = [s for s in known_stations if s]
        
        # Create rows of 5
        rows = []
        chunk_size = 5
        for i in range(0, len(known_stations), chunk_size):
            rows.append(known_stations[i:i + chunk_size])
        
        if not rows:
            rows = [['NO STATIONS MAPPED']]
            
        return rows

    else:
        # Ikes Layout - Dynamic
        known_stations = sorted(list(set(IKES_MAPPING.values())))
        known_stations = [s for s in known_stations if s and s != "NONE"]

        # Create rows of 5
        rows = []
        chunk_size = 5
        for i in range(0, len(known_stations), chunk_size):
            rows.append(known_stations[i:i + chunk_size])

        if not rows:
             rows = [['NO STATIONS MAPPED']]

        return rows


def process_schedule_file(file_path: str, output_dir: str, location: str = "ikes") -> List[str]:
    """Process a single schedule Excel file and generate Manning Charts.

    Returns a list of output file paths created.
    """
    outputs: List[str] = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as exc:
        logging.error(f"Error opening '{file_path}': {exc}")
        return outputs

    ws = wb.active
    header_cell_value = ws['A1'].value or ""
    year_matches = re.findall(r"\d{4}", str(header_cell_value))
    year = year_matches[-1] if year_matches else ""

    header_row = [cell for cell in ws.iter_rows(min_row=2, max_row=2, values_only=True)][0]
    date_columns: List[int] = []
    date_labels: List[str] = []
    for idx, cell in enumerate(header_row):
        if idx == 0 or not cell:
            continue
        m = re.search(r"(\d{2}/\d{2})", str(cell))
        if m:
            date_columns.append(idx)
            label = m.group(1)
            if year and year not in label:
                label = f"{label}/{year}"
            date_labels.append(label)
    if not date_columns:
        logging.warning(f"No valid date columns in '{file_path}'.")
        return outputs

    # Shift definitions based on location
    if location == "southside":
        shifts = [
            {'name': '5.30am-3.00pm', 'meal_periods': 'B L', 'lower': 5.5, 'upper': 15.0},
            {'name': '3.00pm-11.30pm', 'meal_periods': 'D', 'lower': 15.0, 'upper': 23.5},
        ]
    else:
        # Ikes default shifts
        shifts = [
            {'name': '5am-2pm', 'meal_periods': 'B BR', 'lower': 5.0, 'upper': 14.0},
            {'name': '2pm-10pm', 'meal_periods': 'D', 'lower': 14.0, 'upper': 22.0},
            {'name': '10pm-5am', 'meal_periods': 'OVNT', 'lower': 22.0, 'upper': 24.0}, 
        ]

    # Chart row layout
    row_groups = get_stations_layout(location)
    location_name = LOCATIONS.get(location, {}).get("name", location.title())

    generation_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # For each date, produce a workbook
    for file_counter, (col_idx, date_label) in enumerate(zip(date_columns, date_labels), start=1):
        try:
            parsed_date = datetime.strptime(date_label, "%m/%d/%Y")
            weekday_label = f" ({parsed_date.strftime('%A')})"
            date_part = parsed_date.strftime("%a_%d_%b")
        except ValueError:
            parsed_date = None
            weekday_label = ""
            date_part = date_label.replace('/', '-').replace('-', '_')
        
        shift_data = [
            {category: [] for group in row_groups for category in group}
            for _ in shifts
        ]
        
        # Metrics for verification
        total_assignments_found = 0
        mapped_assignments_found = 0
        unmapped_roles_list = set()

        for row in ws.iter_rows(min_row=3, values_only=True):
            role = row[0]
            cell_val = row[col_idx]
            if not cell_val:
                continue
            
            category = get_category(str(role), location)
            
            # Count potential assignments in this cell
            cell_str = str(cell_val).strip()
            assignments = [s for s in re.split(r"\n{2,}", cell_str) if s.strip()]
            
            # Temporary list to hold valid assignments found in this cell
            valid_assignments_in_cell = []

            i = 0
            while i < len(assignments):
                if i + 1 < len(assignments):
                    name = assignments[i].strip()
                    time_range = assignments[i + 1].strip()
                    i += 2
                else:
                    break
                m = re.match(
                    r"(\d{1,2}:\d{2}\s*\w{2})\s*-\s*(\d{1,2}:\d{2}\s*\w{2})",
                    time_range,
                )
                if not m:
                    continue
                
                # Check if valid time parse
                start_time = parse_time(m.group(1))
                if start_time is None:
                    continue
                
                # Check if role is ignored
                if str(role).strip().replace("\n", "").strip().lower() in IGNORED_ROLES:
                    continue

                total_assignments_found += 1
                valid_assignments_in_cell.append((name, m, start_time))

            if not category:
                # If role is not mapped, all assignments in this cell are unmapped
                if valid_assignments_in_cell:
                    unmapped_roles_list.add(str(role))
                continue
            
            # Ensure category exists in layout
            found_layout = False
            for grp in row_groups:
                if category in grp:
                    found_layout = True
                    break
            
            if not found_layout:
                if valid_assignments_in_cell:
                    logging.warning(f"Role '{role}' mapped to '{category}' which is not in the layout.")
                continue

            # Process valid assignments
            for name, m, start_time in valid_assignments_in_cell:
                # Assign to shift based on start time
                shift_index = -1
                
                if location == 'southside':
                    # Simple range check
                    for idx, s in enumerate(shifts):
                        if s['lower'] <= start_time < s['upper']:
                            shift_index = idx
                            break
                        # Handle potential edge case where 11:30pm might be exactly 23.5
                else:
                    # Ikes logic (legacy behavior preservation)
                    if 5 <= start_time < 14:
                        shift_index = 0
                    elif 14 <= start_time < 22:
                        shift_index = 1
                    else:
                        shift_index = 2

                if shift_index != -1:
                    # check if category in that shift's dict (it should be initialized for all)
                    if category in shift_data[shift_index]:
                        entry = f"{name}\n{m.group(1)} - {m.group(2)}"
                        shift_data[shift_index][category].append(entry)
                        mapped_assignments_found += 1
        
        logging.info(f"Verification for {date_label}: Found {total_assignments_found} assignments. Mapped {mapped_assignments_found}.")
        if unmapped_roles_list:
            logging.warning(f"Unmapped roles with assignments: {list(unmapped_roles_list)}")
        if total_assignments_found != mapped_assignments_found:
             logging.warning(f"Mismatch in assignment counts! Missing {total_assignments_found - mapped_assignments_found} assignments.")

        # Create output workbook
        out_wb = openpyxl.Workbook()
        out_wb.remove(out_wb.active)
        for idx_shift, shift_info in enumerate(shifts):
            sheet = out_wb.create_sheet(shift_info['name'])
            # Calculate max columns dynamically
            max_cols = max(len(grp) for grp in row_groups) if row_groups else 1
            if max_cols < 3: max_cols = 3 
            
            from openpyxl.utils import get_column_letter
            end_col_letter = get_column_letter(max_cols)

            sheet.merge_cells(f'A1:{end_col_letter}1')
            sheet['A1'] = f'MANNING CHART - {location_name.upper()}'
            sheet['A1'].font = Font(size=14, bold=True)
            sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            sheet.merge_cells(f'A2:{end_col_letter}2')
            sheet['A2'] = f"Date: {date_label}{weekday_label}    Meal Periods: {shift_info['meal_periods']}    MOD:"
            sheet['A2'].alignment = Alignment(horizontal='left', vertical='center')
            
            for c_idx in range(1, max_cols + 1):
                sheet.column_dimensions[get_column_letter(c_idx)].width = 30
            
            thin = Side(border_style='thin', color='000000')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            row_ptr = 3
            for group in row_groups:
                for col, label in enumerate(group, start=1):
                    hcell = sheet.cell(row=row_ptr, column=col)
                    hcell.value = label
                    hcell.font = Font(bold=True)
                    hcell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    hcell.border = border
                
                data_row = row_ptr + 1
                
                max_lines = 1
                for col, label in enumerate(group, start=1):
                    dcell = sheet.cell(row=data_row, column=col)
                    items = shift_data[idx_shift].get(label, [])
                    cell_text = '\n\n'.join(items) if items else ''
                    dcell.value = cell_text
                    dcell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    dcell.border = border
                    
                    # Calculate lines for this cell
                    # We estimate lines based on newlines and roughly 30 chars per line (column width 30)
                    lines_in_text = cell_text.split('\n')
                    estimated_lines = 0
                    for line in lines_in_text:
                        # Simple wrap estimation: 1 + length // 35
                        estimated_lines += 1 + max(0, (len(line) - 1) // 35)
                    
                    if estimated_lines > max_lines:
                        max_lines = estimated_lines

                sheet.row_dimensions[row_ptr].height = 20
                # Base height approx 15pts per line, minimum 60
                new_height = max(60, max_lines * 15)
                sheet.row_dimensions[data_row].height = new_height
                
                row_ptr += 2

            # Print Settings
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
            sheet.page_setup.fitToPage = True
            sheet.page_setup.fitToHeight = 1
            sheet.page_setup.fitToWidth = 1
                
        out_filename = f"{date_part}_{location_name}_Manning_sheet_{generation_stamp}.xlsx"
        out_path = os.path.join(output_dir, out_filename)
        try:
            out_wb.save(out_path)
            outputs.append(out_filename)
            logging.info(f"Generated '{out_filename}' from '{os.path.basename(file_path)}'.")
        except Exception as exc:
            logging.error(f"Error saving '{out_filename}': {exc}")
    return outputs


######################
# Flask web server   #
######################

app = Flask(__name__, static_folder=STATIC_ROOT, static_url_path="/static")
app.secret_key = os.environ.get("MANNING_APP_SECRET", "manning-standalone-secret")


BASE_CSS = """
/* UI styling lives in static/assets/css. Inline rules reserved for quick overrides. */
.location-toggle {
    margin-bottom: 30px;
    display: flex;
    justify-content: center;
    gap: 0;
}
.location-toggle .btn {
    min-width: 160px;
    border-radius: 0;
    border: 1px solid rgba(255,255,255,0.1);
}
.location-toggle .btn:first-child {
    border-top-left-radius: 30px;
    border-bottom-left-radius: 30px;
}
.location-toggle .btn:last-child {
    border-top-right-radius: 30px;
    border-bottom-right-radius: 30px;
}
.location-toggle .btn-secondary {
    background: transparent;
    color: rgba(255,255,255,0.7);
}
.location-toggle .btn-primary {
    background: #e14eca;
    background-image: linear-gradient(to bottom left, #e14eca, #ba54f5, #e14eca);
    background-size: 210% 210%;
    background-position: top right;
    border-color: transparent;
    box-shadow: 0px 0px 20px 0px rgba(186, 84, 245, 0.5);
}
/* Toast Notifications */
.toast-container {
    position: fixed;
    top: 20px;
    right: 20px;
    z-index: 10000;
    display: flex;
    flex-direction: column;
    gap: 10px;
}
.toast-notification {
    min-width: 300px;
    background: #2b3553;
    color: #ffffff;
    padding: 15px 20px;
    border-radius: 5px;
    box-shadow: 0 4px 6px rgba(0,0,0,0.3);
    border-left: 5px solid #e14eca;
    display: flex;
    align-items: center;
    justify-content: space-between;
    opacity: 0;
    transform: translateX(50px);
    animation: slideIn 0.3s forwards;
    transition: opacity 0.3s ease, transform 0.3s ease;
}
.toast-notification.success {
    border-left-color: #00f2c3; /* Green/Teal for success */
}
.toast-notification.error {
    border-left-color: #fd5d93; /* Red/Pink for error */
}
.toast-notification .close-btn {
    background: none;
    border: none;
    color: rgba(255,255,255,0.6);
    cursor: pointer;
    font-size: 1.2rem;
    line-height: 1;
    margin-left: 10px;
}
.toast-notification .close-btn:hover {
    color: #fff;
}
@keyframes slideIn {
    to {
        opacity: 1;
        transform: translateX(0);
    }
}
.toast-notification.hide {
    opacity: 0;
    transform: translateX(50px);
}
"""


def asset_urls() -> Dict[str, str]:
    """Return URLs for local CSS/JS assets served from /static."""
    import time
    v = int(time.time())
    return {
        "css_black": url_for("static", filename="assets/css/black-dashboard.min.css") + f"?v={v}",
        "css_custom": url_for("static", filename="assets/css/custom.css") + f"?v={v}",
        "css_icons": url_for("static", filename="assets/css/nucleo-icons.css") + f"?v={v}",
        "js_jquery": url_for("static", filename="assets/js/core/jquery.min.js") + f"?v={v}",
        "js_popper": url_for("static", filename="assets/js/core/popper.min.js") + f"?v={v}",
        "js_bootstrap": url_for("static", filename="assets/js/core/bootstrap.min.js") + f"?v={v}",
        "js_black": url_for("static", filename="assets/js/black-dashboard.min.js") + f"?v={v}",
    }


def parse_cell_assignments(cell_value: Optional[str]) -> List[Dict[str, str]]:
    """Split a cell value into individual staff assignments."""
    if not cell_value:
        return []
    text = str(cell_value).strip()
    if not text:
        return []
    blocks = [block.strip() for block in re.split(r"\n{2,}", text) if block.strip()]
    assignments: List[Dict[str, str]] = []
    for block in blocks:
        lines = [line.strip() for line in block.splitlines() if line.strip()]
        if not lines:
            continue
        name = lines[0]
        shift = " ".join(lines[1:]).strip() if len(lines) > 1 else ""
        assignments.append({"name": name, "time": shift})
    return assignments


def build_sheet_structure(ws: "openpyxl.worksheet.worksheet.Worksheet") -> Dict[str, Any]:
    """Create a structured representation of a worksheet's staffing data."""
    rows = list(ws.iter_rows(values_only=True))
    stations: List[Dict[str, Any]] = []
    excel_sections: List[Dict[str, List[str]]] = []
    
    row_idx = 2  # first two rows are header/meta rows
    
    while row_idx < len(rows):
        header_row = rows[row_idx]
        data_row = rows[row_idx + 1] if row_idx + 1 < len(rows) else None
        
        if header_row and data_row is not None:
             # Find how many columns have content in header
            col_count = len(header_row)
            
            headers: List[str] = []
            cells: List[str] = []
            
            has_content = False
            for col_idx in range(col_count):
                header_val = header_row[col_idx]
                station_name = str(header_val).strip() if header_val else ""
                
                # Simple heuristic: stop if several empty headers in a row or use all?
                # We'll rely on the fact that these are generated sheets with borders.
                
                headers.append(station_name)
                
                cell_val = data_row[col_idx] if col_idx < len(data_row) else None
                assignment_text = str(cell_val) if cell_val else ""
                cells.append(assignment_text)
                
                if station_name:
                    has_content = True
                    parsed = parse_cell_assignments(cell_val)
                    stations.append({"station": station_name, "entries": parsed})

            if has_content:
                excel_sections.append({"headers": headers, "cells": cells})
                row_idx += 2
                continue
        
        row_idx += 1
            
    total_entries = sum(len(station["entries"]) for station in stations)
    return {"stations": stations, "total_entries": total_entries, "excel_sections": excel_sections}


def list_output_files() -> List[str]:
    """Return a sorted list of .xlsx files in the output directory."""
    files = [f for f in os.listdir(OUTPUT_DIR) if f.lower().endswith('.xlsx')]
    files.sort()
    return files


@app.route('/', methods=['GET'])
def index() -> str:
    """Render the upload form and list existing outputs."""
    view_mode = request.args.get("view", "current")
    location = request.args.get("location", "ikes") # Default to Ikes
    
    # Validate location
    if location not in LOCATIONS:
        location = "ikes"
    
    show_history = view_mode == "history"
    all_files = list_output_files() if show_history else CURRENT_OUTPUTS.copy()
    
    # Filter files by location
    if location == "southside":
        files = [f for f in all_files if "_Southside_" in f]
    else:
        # Ikes files might explicitly say Ikes or might be legacy (no location name?)
        # Current logic adds location_name to filename: "{date_part}_{location_name}_Manning_sheet..."
        # So Ikes files should have "_Ikes_".
        files = [f for f in all_files if "_Ikes_" in f or ("_Southside_" not in f)]
    
    total_generated = len(CURRENT_OUTPUTS)
    latest_file = CURRENT_OUTPUTS[-1] if CURRENT_OUTPUTS else None
    assets = asset_urls()
    flashes = get_flashed_messages(with_categories=True)
    
    location_name = LOCATIONS[location]["name"]
    title = f"Manning Sheets {location_name}"

    return render_template_string(
        """
<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <title>{{ title }}</title>
    <meta name="viewport" content="width=device-width, initial-scale=1, shrink-to-fit=no">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="{{ css_black }}">
    <link rel="stylesheet" href="{{ css_custom }}">
    <link rel="stylesheet" href="{{ css_icons }}">
    <style>{{ base_css }}</style>
</head>
<body class="app-shell">
<div class="app-surface">
    <!-- Toast Container -->
    <div class="toast-container">
        {% for category, message in flashes %}
        <div class="toast-notification {{ category }}">
            <span>{{ message }}</span>
            <button class="close-btn" onclick="this.parentElement.classList.add('hide'); setTimeout(() => this.parentElement.remove(), 300);">&times;</button>
        </div>
        {% endfor %}
    </div>
    <!-- Sidebar Navigation -->
    <nav class="app-sidebar">
        <div class="brand">
            <i class="tim-icons icon-chart-pie-36"></i> ManningGen
        </div>
        <p class="muted mb-4">Staffing automation for<br><strong>{{ location_name }}</strong></p>
        
        <div class="location-nav">
            <p class="nav-label">Select Location</p>
            <a href="{{ url_for('index', location='ikes', view=view_mode) }}" class="nav-item {{ 'active' if location == 'ikes' else '' }}">
                <i class="tim-icons icon-istanbul"></i> Ikes Dining
            </a>
            <a href="{{ url_for('index', location='southside', view=view_mode) }}" class="nav-item {{ 'active' if location == 'southside' else '' }}">
                <i class="tim-icons icon-bank"></i> Southside
            </a>
        </div>

        <div class="sidebar-footer mt-auto">
             <p class="small text-muted text-center">v2.2 Nano Banana</p>
        </div>
    </nav>

    <!-- Main Content -->
    <main class="app-content">
        <div class="container-fluid p-0">
            <!-- Header Row -->
            <div class="row mb-4 align-items-center">
                <div class="col-12">
                     <h2 class="page-title">{{ title }}</h2>
                     <p class="text-muted">Generate compliant manning charts from MyStaff schedules.</p>
                </div>
            </div>

            <div class="row g-4 mb-5">
                <!-- Usage Instructions Card -->
                <div class="col-12 col-xl-5">
                    <div class="app-card h-100">
                        <h4 class="mb-3">Instructions</h4>
                        <ol class="instruction-list ps-3">
                            <li class="mb-2">Visit <strong>MyStaff</strong> and select the weekly schedule and click on print.</li>
                            <li class="mb-2">Switch view to <strong>Task</strong> (top right corner).</li>
                            <li class="mb-2">Click on <strong>Print</strong> (or the Excel export button) to download the file.</li>
                            <li class="mb-2">Upload the file here and click <strong>Generate Charts</strong>.</li>
                            <li>Your charts will appear below.</li>
                        </ol>
                    </div>
                </div>

                <!-- Upload Card -->
                <div class="col-12 col-xl-7">
                     <div class="app-card h-100 hero-gradient">
                        <div class="d-flex align-items-center justify-content-between mb-4">
                            <div>
                                <h3 class="mb-1 text-white">Create New Manning Sheets</h3>
                                <p class="opacity-75 mb-0">Upload MyStaff export (.xlsx)</p>
                            </div>
                            <div class="icon-shape bg-white text-primary rounded-circle shadow-sm d-flex align-items-center justify-content-center" style="width: 40px; height: 40px;">
                                <i class="tim-icons icon-cloud-upload-94" style="font-size: 1.2rem;"></i>
                            </div>
                        </div>
                          
                        <form action="{{ url_for('upload') }}" method="post" enctype="multipart/form-data" class="upload-form">
                            <input type="hidden" name="location" value="{{ location }}">
                            <div class="file-drop-area w-100 mb-3">
                                <span class="choose-file-btn mb-2">Choose File</span>
                                <span class="file-msg small text-white-50">or drag and drop file here</span>
                                <input class="file-input" type="file" name="file" accept=".xlsx" required>
                            </div>
                            <button type="submit" class="btn btn-white w-100 btn-lg fw-bold">Generate Charts &rarr;</button>
                        </form>
                     </div>
                </div>
            </div>

            <!-- Stats Row -->
            <div class="row g-4 mb-4">
                 <div class="col-6 col-md-3">
                    <div class="app-card text-center p-3">
                        <h2 class="mb-0 text-primary">{{ total_generated }}</h2>
                        <small class="text-muted text-uppercase">Total Charts</small>
                    </div>
                 </div>
                 <div class="col-6 col-md-9">
                    <div class="app-card p-3 d-flex align-items-center justify-content-between position-relative overflow-hidden">
                        <div style="min-width: 0;">
                             <small class="text-muted text-uppercase d-block">Latest Workbook</small>
                             <span class="text-white text-truncate d-block" style="max-width: 100%;">{{ latest_file or "No files yet" }}</span>
                        </div>
                        <div class="icon-shape bg-primary text-white rounded-circle shadow-sm flex-shrink-0 ms-3 d-flex align-items-center justify-content-center" style="width:40px;height:40px;">
                            <i class="tim-icons icon-calendar-60" style="font-size: 1rem;"></i>
                        </div>
                    </div>
                 </div>
            </div>

            <!-- History Section -->
            <div class="row">
                <div class="col-12">
                    <div class="app-card">
                        <div class="d-flex justify-content-between align-items-center mb-4 flex-wrap gap-2">
                             <h4 class="mb-0">{{ "History Archive" if show_history else "Recent Sessions" }}</h4>
                             <div class="btn-group">
                                 <a href="{{ url_for('index', view='current', location=location) }}" class="btn btn-sm btn-{{ 'primary' if not show_history else 'simple' }}">Current</a>
                                 <a href="{{ url_for('index', view='history', location=location) }}" class="btn btn-sm btn-{{ 'primary' if show_history else 'simple' }}">History</a>
                             </div>
                        </div>
                        
                        {% if files %}
                        <div class="table-responsive">
                            <table class="table tablesorter align-middle" id="">
                                <thead class="text-primary">
                                    <tr>
                                        <th>Generated File</th>
                                        <th class="text-right">Actions</th>
                                    </tr>
                                </thead>
                                <tbody>
                                    {% for fname in files %}
                                    <tr>
                                        <td>
                                            <div class="d-flex align-items-center gap-2">
                                                <i class="tim-icons icon-single-copy-04 text-muted"></i>
                                                <span class="fw-bold">{{ fname }}</span>
                                            </div>
                                        </td>
                                        <td class="text-right">
                                            <div class="btn-group">
                                                <a href="{{ url_for('view_file', filename=fname) }}" class="btn btn-sm btn-info">
                                                    <i class="tim-icons icon-zoom-split"></i> View
                                                </a>
                                                <a href="{{ url_for('download_file', filename=fname) }}" class="btn btn-sm btn-success">
                                                    <i class="tim-icons icon-cloud-download-93"></i> Download
                                                </a>
                                                <a href="{{ url_for('view_file', filename=fname) }}?print=true" target="_blank" class="btn btn-sm btn-warning">
                                                    <i class="tim-icons icon-print"></i> Print
                                                </a>
                                            </div>
                                        </td>
                                    </tr>
                                    {% endfor %}
                                </tbody>
                            </table>
                        </div>
                        {% else %}
                        <div class="text-center py-5">
                            <h5 class="text-muted">No charts found</h5>
                        </div>
                        {% endif %}
                    </div>
                </div>
            </div>
            
        </div>
    </main>
</div>

<script src="{{ js_jquery }}"></script>
<script src="{{ js_popper }}"></script>
<script src="{{ js_bootstrap }}"></script>
<script src="{{ js_black }}"></script>
<script>
    $('.file-input').on('change', function() {
      var filesCount = $(this)[0].files.length;
      var textContainer = $(this).prev();
      if (filesCount === 1) {
        textContainer.text($(this).val().split('\\\\').pop());
      } else {
        textContainer.text('or drag and drop file here');
      }
    });
</script>
</body>
</html>
        """,
        files=files,
        total_generated=total_generated,
        latest_file=latest_file,
        base_css=BASE_CSS,
        flashes=flashes,
        show_history=show_history,
        title=title,
        location=location,
        location_name=location_name,
        view_mode=view_mode,
        **assets,
    )


@app.route('/upload', methods=['POST'])
def upload() -> str:
    """Handle file upload, save to input directory, process it, and redirect."""
    global CURRENT_OUTPUTS
    
    location = request.form.get("location", "ikes")
    
    if 'file' not in request.files:
        flash("No file selected.", "error")
        return redirect(url_for('index', location=location))
    
    file = request.files['file']
    if file.filename == '':
        flash("Please choose a file before uploading.", "error")
        return redirect(url_for('index', location=location))
    if not file.filename.lower().endswith('.xlsx'):
        flash("Only .xlsx files are supported. Upload the MyStaff shift schedule exported in Task Wise view (.xlsx).", "error")
        return redirect(url_for('index', location=location))
    # Save the uploaded file with a timestamp to avoid collisions
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_filename = f"{timestamp}_{os.path.basename(file.filename)}"
    input_path = os.path.join(INPUT_DIR, safe_filename)
    file.save(input_path)
    logging.info(f"Uploaded schedule saved to '{input_path}'.")
    
    # Validate location
    if not validate_file_location(input_path, location):
        # Remove the invalid file
        try:
            os.remove(input_path)
        except OSError:
            pass
            
        location_name = LOCATIONS.get(location, {}).get("name", location)
        flash(f'Please upload "{location_name}" schedule by following the instructions', "error")
        return redirect(url_for('index', location=location))

    # Process the uploaded schedule
    outputs = process_schedule_file(input_path, OUTPUT_DIR, location=location)
    if outputs:
        logging.info(f"Generated {len(outputs)} output file(s) from '{safe_filename}'.")
        CURRENT_OUTPUTS = outputs
        flash(f"Successfully generated {len(outputs)} chart(s).", "success")
    else:
        logging.warning(f"No output files generated from '{safe_filename}'.")
        flash("Unable to process that file. Upload the MyStaff shift schedule exported in Task Wise view (.xlsx) and try again.", "error")
    # Redirect to index to show new files
    return redirect(url_for('index', location=location))


@app.route('/download/<path:filename>')
def download_file(filename: str):
    """Serve a file from the output directory."""
    return send_from_directory(OUTPUT_DIR, filename, as_attachment=True)


@app.route('/view/<path:filename>')
def view_file(filename: str):
    """Render an HTML representation of a generated workbook."""
    file_path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(file_path) or not filename.lower().endswith('.xlsx'):
        return abort(404)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as exc:
        logging.error(f"Error reading '{file_path}': {exc}")
        return f"Error reading workbook: {exc}", 500

    sheet_tables: List[Dict[str, Any]] = []
    
    # Attempt to deduce location from filename for the Title
    location_title = "Manning Sheets"
    if "_Ikes_" in filename:
        location_title = "Manning Sheets - Ikes Dining"
    elif "_Southside_" in filename:
        location_title = "Manning Sheets - Southside"

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = build_sheet_structure(ws)
        
        # Extract metadata from A2 if available
        header_meta = ws['A2'].value if ws['A2'].value else ""
        
        sheet_tables.append(
            {
                "name": sheet_name,
                "stations": sheet_data["stations"],
                "total_entries": sheet_data["total_entries"],
                "excel_sections": sheet_data["excel_sections"],
                "header_metadata": header_meta
            }
        )

    assets = asset_urls()
    return render_template_string(
        """
<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <title>{{ filename }} - Viewer</title>
    <meta name="viewport" content="width=device-width, initial-scale=1, shrink-to-fit=no">
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="{{ css_black }}">
    <link rel="stylesheet" href="{{ css_custom }}">
    <link rel="stylesheet" href="{{ css_icons }}">
    <style>{{ base_css }}</style>
</head>
<body class="app-shell viewer-shell">
<div class="app-surface">
    <!-- Sidebar Navigation (Same as Index) -->
    <nav class="app-sidebar">
        <div class="brand">
            <i class="tim-icons icon-chart-pie-36"></i> ManningGen
        </div>
        <p class="muted mb-4">Staffing automation for<br><strong>Ikes/Southside</strong></p>
        
        <div class="location-nav">
             <a href="{{ url_for('index', location='ikes') }}" class="nav-item">
                <i class="tim-icons icon-istanbul"></i> Ikes Dining
            </a>
            <a href="{{ url_for('index', location='southside') }}" class="nav-item">
                <i class="tim-icons icon-bank"></i> Southside
            </a>
        </div>
        
        <div class="mt-4 px-2">
            <a href="{{ url_for('index') }}" class="btn btn-sm btn-simple text-white border-white w-100">
                <i class="tim-icons icon-minimal-left"></i> Back to Dashboard
            </a>
        </div>

        <div class="sidebar-footer mt-auto">
             <p class="small text-muted text-center">v2.2 Nano Banana</p>
        </div>
    </nav>

    <!-- Main Content -->
    <main class="app-content">
        <div class="container-fluid p-0">
             <!-- Viewer Header -->
             <div class="viewer-actions p-4 mb-4 rounded d-flex justify-content-between align-items-center bg-dark shadow-sm d-print-none">
                <div class="d-flex align-items-center gap-3">
                     <div class="icon-shape bg-info text-white rounded-circle shadow-sm d-flex align-items-center justify-content-center flex-shrink-0" style="width: 40px; height: 40px;">
                        <i class="tim-icons icon-paper"></i>
                     </div>
                     <div>
                         <h4 class="mb-0 text-white">{{ filename }}</h4>
                         <small class="text-muted">Viewing generated workbook</small>
                     </div>
                </div>
                <div>
                     <a href="{{ url_for('download_file', filename=filename) }}" class="btn btn-success btn-sm me-2"><i class="tim-icons icon-cloud-download-93"></i> Download</a>
                     <button onclick="handlePrint()" class="btn btn-warning btn-sm"><i class="tim-icons icon-print"></i> Print All Shifts</button>
                </div>
            </div>

            <!-- Tabs -->
            <div class="viewer-tabs mb-4 text-center d-print-none">
                {% for table in sheets %}
                <button class="shift-btn {{ 'active' if loop.first else '' }}" onclick="showSheet('{{ table.name }}', this)">
                    {{ table.name }}
                </button>
                {% endfor %}
            </div>

            <!-- Sheets -->
            {% for table in sheets %}
            <div id="sheet-{{ table.name }}" class="shift-panel {{ 'active' if loop.first else '' }}">
                <div class="sheet-card">
                     <!-- Print Header -->
                     <div class="print-header d-none d-print-block mb-3 text-center border-bottom border-dark pb-2">
                         <h2 class="mb-1">{{ location_title }}</h2>
                         <h3 class="mb-1">{{ table.name }}</h3>
                         <p class="mb-0 text-muted small" style="white-space: pre-wrap;">{{ table.header_metadata }}</p>
                     </div>
                     
                    <div class="excel-view">
                        <div class="table-responsive">
                            <table class="excel-table">
                                {% for block in table.excel_sections %}
                                    <tr>
                                        {% for header in block.headers %}
                                            <th>{{ header }}</th>
                                        {% endfor %}
                                    </tr>
                                    <tr>
                                        {% for cell in block.cells %}
                                            <td>{% if cell %}{{ cell.replace('\\n', '<br>')|safe }}{% else %}&nbsp;{% endif %}</td>
                                        {% endfor %}
                                    </tr>
                                {% endfor %}
                            </table>
                        </div>
                    </div>
                </div>
            </div>
            {% endfor %}
        </div>
    </main>
</div>

<script src="{{ js_jquery }}"></script>
<script src="{{ js_popper }}"></script>
<script src="{{ js_bootstrap }}"></script>
<script src="{{ js_black }}"></script>
<script>
function showSheet(name, btn) {
    document.querySelectorAll('.shift-panel').forEach(el => el.classList.remove('active'));
    document.querySelectorAll('.shift-btn').forEach(el => el.classList.remove('active'));
    document.getElementById('sheet-' + name).classList.add('active');
    btn.classList.add('active');
}

function handlePrint() {
    // Auto-scale to fit landscape page (approx 1000px safe width)
    const MAX_WIDTH = 1050; 
    let maxTableWidth = 0;
    
    // Find widest table
    document.querySelectorAll('.excel-table').forEach(tbl => {
        if (tbl.offsetWidth > maxTableWidth) maxTableWidth = tbl.offsetWidth;
    });

    if (maxTableWidth > MAX_WIDTH) {
        const scale = MAX_WIDTH / maxTableWidth;
        document.body.style.zoom = scale;
    } else {
        document.body.style.zoom = 1;
    }

    // Small delay to allow render
    // Small delay to allow render
    setTimeout(() => {
        // Reset zoom after print dialog closes using onafterprint
        window.onafterprint = function() {
            document.body.style.zoom = 1;
        };
        window.print();
        
        // Fallback for browsers that might not fire onafterprint reliably or if blocked
        window.addEventListener('focus', function() {
             document.body.style.zoom = 1;
        }, { once: true });
    }, 100);
}

const urlParams = new URLSearchParams(window.location.search);
if (urlParams.get('print')) {
    handlePrint();
}

// Auto-dismiss toasts
document.addEventListener('DOMContentLoaded', () => {
    const toasts = document.querySelectorAll('.toast-notification');
    toasts.forEach(toast => {
        setTimeout(() => {
            toast.classList.add('hide');
            setTimeout(() => toast.remove(), 300);
        }, 5000); // 5 seconds
    });
});
</script>
</body>
</html>
        """,
        filename=filename,
        sheets=sheet_tables,
        location_title=location_title,
        base_css=BASE_CSS,
        **assets,
    )


@app.route('/view_log')
def view_log() -> str:
    """Display the log file in the browser."""
    try:
        with open(log_filename, "r", encoding="utf-8") as f:
            content = f.read()
    except Exception as exc:
        content = f"Error reading log: {exc}"
    return render_template_string(
        """
<!doctype html>
<html>
<head>
    <title>System Log</title>
    <style>body{background:#1e1e2f;color:#e1e4e8;font-family:monospace;padding:20px;white-space:pre-wrap;}</style>
</head>
<body>{{ content }}</body>
</html>
        """,
        content=content,
    )


if __name__ == '__main__':
    open_browser = True
    if len(sys.argv) > 1 and sys.argv[1] == '--no-browser':
        open_browser = False

    # Start the server
    port = int(os.environ.get("PORT", 5000))
    if open_browser:
        threading.Timer(1.0, lambda: webbrowser.open(f"http://127.0.0.1:{port}")).start()
    
    app.run(host='0.0.0.0', port=port)
